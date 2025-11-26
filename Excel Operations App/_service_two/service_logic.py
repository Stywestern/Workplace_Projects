# ==========================================
# 1. IMPORTS
# ==========================================

# Third-Party Libraries
import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from _config.settings import load_config
from _utils.functions import open_file

# Initialize Logger
logger = logging.getLogger('AppLogger')


# ==========================================
# 2. COMPARISON LOGIC
# ==========================================

def compare() -> None:
    """
    Compares two Excel files based on a unique key.
    Outputs a new Excel file highlighting changes (Old -> New).
    """
    logger.info("Starting File Comparison")
    print("Loading configuration...")
    
    config = load_config()

    # Load Paths & Configs
    file1_path = config["service_two"]["file1"]["file_path"]
    sheet1 = config["service_two"]["file1"]["sheet_name"]

    file2_path = config["service_two"]["file2"]["file_path"]
    sheet2 = config["service_two"]["file2"]["sheet_name"]

    unique_key = config["service_two"]["unique_id_column"]
    compare_columns = config["service_two"]["compare_columns"]
    
    # Ensure unique_key is not in compare_columns (avoid redundancy)
    if unique_key in compare_columns:
        compare_columns.remove(unique_key)

    try:
        # Load DataFrames
        print("Reading Excel files...")
        df1 = pd.read_excel(file1_path, sheet_name=sheet1, engine='openpyxl')
        df2 = pd.read_excel(file2_path, sheet_name=sheet2, engine='openpyxl')

        # ==========================================
        # ALIGNMENT & INTERSECTION
        # ==========================================
        
        # 1. Align Columns
        common_columns = list(set(df1.columns).intersection(set(df2.columns)))
        if unique_key not in common_columns:
            logger.error(f"Unique key '{unique_key}' not found in both files.")
            print(f"❌ Unique key '{unique_key}' not found in both files.")
            return

        # Filter columns
        compare_columns = [col for col in compare_columns if col in common_columns]
        stable_columns = [col for col in common_columns if col not in compare_columns]

        # Ensure Unique Key is first
        if unique_key in stable_columns:
            stable_columns.remove(unique_key)
        stable_columns.insert(0, unique_key)

        reordered_columns = stable_columns + compare_columns
        df1 = df1[reordered_columns].copy()
        df2 = df2[reordered_columns].copy()

        # 2. Align Rows (Intersection of Keys)
        df1_keys = set(df1[unique_key].dropna())
        df2_keys = set(df2[unique_key].dropna())
        common_keys = df1_keys.intersection(df2_keys)

        df1_aligned = df1[df1[unique_key].isin(common_keys)].copy()
        df2_aligned = df2[df2[unique_key].isin(common_keys)].copy()

        print(f"Aligned DataFrames: {len(df1_aligned)} shared rows found.")

        # 3. Sort (Crucial for row-by-row comparison)
        sortable_columns = [col for col in df1.columns if col not in compare_columns]
        if unique_key in sortable_columns:
            sortable_columns.remove(unique_key)
        sortable_columns.insert(0, unique_key)

        df1_aligned = df1_aligned.sort_values(by=sortable_columns).reset_index(drop=True)
        df2_aligned = df2_aligned.sort_values(by=sortable_columns).reset_index(drop=True)

        # ==========================================
        # VALUE COMPARISON
        # ==========================================
        
        print("Comparing values...")
        result_df = df1_aligned.copy()
        mismatches = [] # Stores (row_index, col_name)

        # Iterate rows
        for idx in range(len(df1_aligned)):
            for col in compare_columns:
                val1 = df1_aligned.iloc[idx][col]
                val2 = df2_aligned.iloc[idx][col]

                # Treat NaNs as equal
                if pd.isna(val1) and pd.isna(val2):
                    continue

                if val1 != val2:
                    # Record change string
                    result_df.at[idx, col] = f"{val1} → {val2}"
                    mismatches.append((idx, col))

        print(f"{len(mismatches)} individual cell changes found.")

        # ==========================================
        # EXCEL GENERATION
        # ==========================================
        
        print("Generating report...")
        wb = Workbook()

        # --- Sheet 1: Info ---
        info_sheet = wb.active
        info_sheet.title = "Info"
        info_sheet["A1"], info_sheet["B1"] = "File 1 Path:", file1_path
        info_sheet["A2"], info_sheet["B2"] = "File 2 Path:", file2_path
        info_sheet["A3"], info_sheet["B3"] = "Explanation:", "'Old → New'"

        # --- Sheet 2: Full Result ---
        wb.create_sheet("Full Result")
        full_sheet = wb["Full Result"]

        # Write Headers
        for col_idx, col_name in enumerate(result_df.columns, 1):
            full_sheet.cell(row=1, column=col_idx, value=col_name)

        # Write Data
        for r_idx, row in result_df.iterrows():
            for c_idx, value in enumerate(row, 1):
                full_sheet.cell(row=r_idx + 2, column=c_idx, value=value)

        # Highlight Mismatches
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Set of row indices (0-based) that have changes
        highlighted_row_indices = set()

        for row_idx, col_name in mismatches:
            # Excel is 1-based, plus 1 for header = row_idx + 2
            excel_row = row_idx + 2
            # Find column index
            col_idx = list(result_df.columns).index(col_name) + 1
            
            # Apply fill
            full_sheet.cell(row=excel_row, column=col_idx).fill = yellow_fill
            
            # Track this row for the Filtered Sheet
            highlighted_row_indices.add(row_idx)

        # --- Sheet 3: Filtered Result (Optimized) ---
        wb.create_sheet("Filtered Result")
        filtered_sheet = wb["Filtered Result"]

        # Write Headers
        for col_idx, col_name in enumerate(result_df.columns, 1):
            filtered_sheet.cell(row=1, column=col_idx, value=col_name)

        # Write only changed rows sequentially
        current_write_row = 2
        
        # Iterate through the Original Result DF
        for r_idx, row in result_df.iterrows():
            if r_idx in highlighted_row_indices:
                # Write the whole row
                for c_idx, value in enumerate(row, 1):
                    cell = filtered_sheet.cell(row=current_write_row, column=c_idx, value=value)
                    
                    # Re-apply highlight if this specific cell changed
                    # (Check if this (r_idx, col_name) exists in mismatches)
                    col_name = result_df.columns[c_idx - 1]
                    if (r_idx, col_name) in mismatches:
                         cell.fill = yellow_fill
                
                current_write_row += 1

        # ==========================================
        # SAVE
        # ==========================================
        
        output_dir = "outputs/compare_changes"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        file1_name = os.path.basename(file1_path).rsplit('.', 1)[0]
        file2_name = os.path.basename(file2_path).rsplit('.', 1)[0]
        output_path = f"{output_dir}/comparison_{file1_name}_vs_{file2_name}.xlsx"

        wb.save(output_path)
        print(f"\n✅ Comparison complete. Saved to: {output_path}")
        
        open_file(output_path)

    except Exception as e:
        logger.error(f"Comparison failed: {e}")
        print(f"An error occurred: {e}")
        